#!/usr/bin/env python3
"""Generate Electro-V2 use case specification Excel for diagram validation."""

import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "Electro-V2-Use-Case.xlsx"
OUTPUT_ALT = Path(__file__).resolve().parent.parent / "docs" / "Electro-V2-Use-Case-matrix.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D10024")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2B2D42")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USE_CASES = [
    ("UC-01", "Login", "Authenticate with email and password; receive API token.", "Admin, User"),
    ("UC-02", "Verify Input", "Validate email/password fields before login (included in Login).", "— (included)"),
    ("UC-03", "Sign in with Google", "OAuth via Google; extends Login (login page only).", "User"),
    ("UC-04", "Display Login error", "Show error when credentials are wrong or email unverified.", "— (extend)"),
    ("UC-05", "Sign up", "Register customer account (name, email, phone, password).", "Visitor"),
    ("UC-06", "Verify Email", "Send / confirm email verification after sign up.", "— (extend)"),
    ("UC-07", "Sign up with Google", "Optional: Google may create account via Login flow — not a separate signup button.", "— (extend only)"),
    ("UC-08", "Logout", "End session; revoke token (with confirmation dialog).", "Admin, User"),
    ("UC-09", "Reset Password", "Forgot password + reset via email link (must already have an account).", "User, Admin"),
    ("UC-10", "Profile Management", "View/update profile (name, phone) and change password.", "Admin, User"),
    ("UC-11", "CRUD Categories", "Admin creates, reads, updates, deletes product categories.", "Admin"),
    ("UC-12", "CRUD Brands", "Admin manages brands.", "Admin"),
    ("UC-13", "CRUD Products", "Admin manages products (name, price, stock, descriptions).", "Admin"),
    ("UC-14", "Manage Product Image", "Assign/sync product image paths (included in CRUD Products).", "— (included)"),
    ("UC-15", "CRUD Cities", "Admin manages delivery cities.", "Admin"),
    ("UC-16", "CRUD Areas", "Admin manages areas within cities (included in CRUD Cities).", "— (included)"),
    ("UC-17", "Orders Management", "Admin lists all orders and views details.", "Admin"),
    ("UC-18", "Update Order Status", "Admin changes order status (pending/paid/shipped/etc.).", "— (included)"),
    ("UC-19", "Manage Exchange Rate", "Admin sets USD → SP rate for display.", "Admin"),
    ("UC-20", "Browse Store", "Browse/filter/search products in the store catalog.", "Visitor, User"),
    ("UC-21", "View Products", "View product listing (store/home widgets).", "Visitor, User"),
    ("UC-22", "View Product Details", "Open product page (images, price, description, reviews).", "Visitor, User"),
    ("UC-23", "Manage Cart", "Add/update/remove cart items (requires login).", "User"),
    ("UC-24", "Place Order", "Checkout: create order from cart.", "User"),
    ("UC-25", "Select City and Area", "Choose delivery city/area at checkout (included in Place Order).", "— (included)"),
    ("UC-26", "View my Orders", "Customer views own order history and details.", "User"),
    ("UC-27", "Submit Product Review", "Authenticated user posts rating/comment on a product.", "User"),
]

INCLUDES = [
    ("Login", "Verify Input", "Login always validates input before authenticating."),
    ("CRUD Products", "Manage Product Image", "Product create/update sets image path."),
    ("CRUD Cities", "CRUD Areas", "Areas belong to cities; managed together in admin."),
    ("Orders Management", "Update Order Status", "Status change is part of managing orders."),
    ("Place Order", "Select City and Area", "Checkout requires city_id and area_id."),
]

EXTENDS = [
    ("Sign in with Google", "Login", "Alternative path to Login (login page button)."),
    ("Display Login error", "Login", "Only when login fails."),
    ("Verify Email", "Sign up", "Triggered after registration; arrow points TO Sign up."),
    ("Sign up with Google", "Sign up", "Optional: Google auto-registration. Not a separate signup-page button in Electro-V2."),
]

ACTORS = ["Admin", "User", "Visitor"]

# Y = primary actor link in diagram
ACTOR_MATRIX = {
    "Login": {"Admin": "Y", "User": "Y", "Visitor": ""},
    "Verify Input": {"Admin": "", "User": "", "Visitor": ""},
    "Sign in with Google": {"Admin": "", "User": "Y", "Visitor": ""},
    "Display Login error": {"Admin": "", "User": "", "Visitor": ""},
    "Sign up": {"Admin": "", "User": "", "Visitor": "Y"},
    "Verify Email": {"Admin": "", "User": "", "Visitor": ""},
    "Sign up with Google": {"Admin": "", "User": "", "Visitor": ""},
    "Logout": {"Admin": "Y", "User": "Y", "Visitor": ""},
    "Reset Password": {"Admin": "Y", "User": "Y", "Visitor": ""},
    "Profile Management": {"Admin": "Y", "User": "Y", "Visitor": ""},
    "CRUD Categories": {"Admin": "Y", "User": "", "Visitor": ""},
    "CRUD Brands": {"Admin": "Y", "User": "", "Visitor": ""},
    "CRUD Products": {"Admin": "Y", "User": "", "Visitor": ""},
    "Manage Product Image": {"Admin": "", "User": "", "Visitor": ""},
    "CRUD Cities": {"Admin": "Y", "User": "", "Visitor": ""},
    "CRUD Areas": {"Admin": "", "User": "", "Visitor": ""},
    "Orders Management": {"Admin": "Y", "User": "", "Visitor": ""},
    "Update Order Status": {"Admin": "", "User": "", "Visitor": ""},
    "Manage Exchange Rate": {"Admin": "Y", "User": "", "Visitor": ""},
    "Browse Store": {"Admin": "", "User": "Y", "Visitor": "Y"},
    "View Products": {"Admin": "", "User": "Y", "Visitor": "Y"},
    "View Product Details": {"Admin": "", "User": "Y", "Visitor": "Y"},
    "Manage Cart": {"Admin": "", "User": "Y", "Visitor": ""},
    "Place Order": {"Admin": "", "User": "Y", "Visitor": ""},
    "Select City and Area": {"Admin": "", "User": "", "Visitor": ""},
    "View my Orders": {"Admin": "", "User": "Y", "Visitor": ""},
    "Submit Product Review": {"Admin": "", "User": "Y", "Visitor": ""},
}

NOTES = [
    "ACTOR DEFINITIONS: Visitor = not logged in, no account yet (browse + sign up only). User = registered customer. Admin = shop administrator.",
    "Visitor does NOT Login, Reset Password, or Sign in with Google. Those belong to User/Admin (people who already have an account).",
    "Why? Reset Password is for account holders who forgot their password — not for a guest with no account. Login is for User/Admin, not Visitor.",
    "After Sign up, the person becomes a User. Sign in with Google extends Login — link User (and Admin if needed) to Login only.",
    "CORRECT «include» relations: Login→Verify Input; CRUD Products→Manage Product Image; CRUD Cities→CRUD Areas; Orders Management→Update Order Status; Place Order→Select City and Area.",
    "CORRECT «extend» (arrow from extension TO base): Sign in with Google→Login; Display Login error→Login; Verify Email→Sign up.",
    "Do NOT link Visitor to: Login, Logout, Reset Password, Manage Cart, Place Order, View my Orders, Submit Product Review.",
    "Do link Visitor to: Sign up, Browse Store, View Products, View Product Details.",
    "Do link Admin to: Login, Logout, Profile Management, all CRUD use cases, Orders Management, Manage Exchange Rate.",
    "After login redirect: Admin → Dashboard; User → Home.",
]


def style_header_row(ws, row, ncol):
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def write_use_cases(ws):
    ws.title = "Use Cases"
    headers = ["ID", "Use Case", "Description", "Actors"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for row in USE_CASES:
        ws.append(list(row))
        for c in range(1, 5):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP
            ws.cell(row=ws.max_row, column=c).border = BORDER
    autosize(ws)


def write_includes(ws):
    ws.append(["Base Use Case", "«include»", "Included Use Case", "Notes"])
    style_header_row(ws, 1, 4)
    for base, inc, note in INCLUDES:
        ws.append([base, "include", inc, note])
        for c in range(1, 5):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP
            ws.cell(row=ws.max_row, column=c).border = BORDER
    autosize(ws)


def write_extends(ws):
    ws.append(["Extension Use Case", "«extend»", "Base Use Case", "UML: dashed arrow points TO base"])
    style_header_row(ws, 1, 4)
    for ext, base, note in EXTENDS:
        ws.append([ext, "extend", base, note])
        for c in range(1, 5):
            ws.cell(row=ws.max_row, column=c).alignment = WRAP
            ws.cell(row=ws.max_row, column=c).border = BORDER
    autosize(ws)


def write_actor_matrix(ws):
    main_uc = list(ACTOR_MATRIX.keys())
    ws.append(["Use Case \\ Actor"] + ACTORS)
    style_header_row(ws, 1, len(ACTORS) + 1)
    for uc in main_uc:
        row = [uc] + [ACTOR_MATRIX.get(uc, {}).get(a, "") for a in ACTORS]
        ws.append(row)
        for c in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=ws.max_row, column=c).border = BORDER
    ws.column_dimensions["A"].width = 28
    for i in range(2, 5):
        ws.column_dimensions[get_column_letter(i)].width = 12


def write_notes(ws):
    ws.append(["Diagram review notes (Electro-V2)"])
    ws.cell(1, 1).fill = SECTION_FILL
    ws.cell(1, 1).font = SECTION_FONT
    ws.merge_cells("A1:D1")
    row = 2
    for note in NOTES:
        ws.cell(row=row, column=1, value=note)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).alignment = WRAP
        row += 1
    ws.column_dimensions["A"].width = 100


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    write_use_cases(wb.active)
    ws_inc = wb.create_sheet("Include Relations")
    write_includes(ws_inc)
    ws_ext = wb.create_sheet("Extend Relations")
    write_extends(ws_ext)
    ws_mat = wb.create_sheet("Actor Matrix")
    write_actor_matrix(ws_mat)
    ws_notes = wb.create_sheet("Review Notes")
    write_notes(ws_notes)
    target = OUTPUT
    try:
        wb.save(target)
    except PermissionError:
        target = OUTPUT_ALT
        wb.save(target)
        print(f"Note: {OUTPUT.name} is open — saved to alternate file.")
    print(f"Created: {target}")


if __name__ == "__main__":
    main()
