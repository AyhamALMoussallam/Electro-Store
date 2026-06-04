#!/usr/bin/env python3
"""Patch edited ERD Excel: replace description with description_en / description_ar."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parent.parent / "docs"
FILES = [
    ROOT / "Electro-V2-ERD-columns.xlsx",
    ROOT / "Electro-V2-ERD.xlsx",
]

FIELD_FONT = Font(size=10)
ALIGN = Alignment(wrap_text=True, vertical="top")
DESC_EN = "description_en (text) — English; shown when site language is EN"
DESC_AR = "description_ar (text) — Arabic; shown when site language is AR"


def find_product_column(ws):
    for col in range(1, ws.max_column + 1):
        title = ws.cell(1, col).value
        if title and str(title).strip().lower() in ("product", "products"):
            return col
    return 4


def write_field(ws, row, col, text):
    cell = ws.cell(row, col, value=text)
    cell.font = FIELD_FONT
    cell.alignment = ALIGN


def patch_products_column(ws, col):
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, col).value
        if not val:
            continue
        text = str(val).strip()
        if text.startswith("description_en ("):
            write_field(ws, row, col, DESC_EN)
            if row + 1 <= ws.max_row:
                nxt = ws.cell(row + 1, col).value
                if nxt and str(nxt).startswith("description_ar ("):
                    write_field(ws, row + 1, col, DESC_AR)
            return True
        if text.startswith("description (") or text == "description (text)":
            shifted = []
            r = row + 1
            while r <= ws.max_row:
                v = ws.cell(r, col).value
                if not v:
                    break
                s = str(v)
                if s.startswith(("RELATIONSHIP", "Parent (", "categories", "brands")):
                    break
                if "FK column" in s or s in ("CASCADE", "SET NULL"):
                    break
                shifted.append(v)
                r += 1
            write_field(ws, row, col, DESC_EN)
            ws.insert_rows(row + 1)
            write_field(ws, row + 1, col, DESC_AR)
            for i, old in enumerate(shifted):
                write_field(ws, row + 2 + i, col, old)
            return True
    return False


def main():
    for path in FILES:
        if not path.exists():
            print(f"Skip (missing): {path}")
            continue
        wb = load_workbook(path)
        ws = wb.active
        col = find_product_column(ws)
        if patch_products_column(ws, col):
            wb.save(path)
            print(f"Updated: {path}")
        else:
            print(f"No description field in: {path}")
        wb.close()


if __name__ == "__main__":
    main()
