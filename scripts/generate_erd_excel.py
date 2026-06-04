#!/usr/bin/env python3
"""Generate Electro-V2 ERD as a single-sheet Excel workbook (one column per table)."""

import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "docs" / "Electro-V2-ERD.xlsx"
OUTPUT_ALT = Path(__file__).resolve().parent.parent / "docs" / "Electro-V2-ERD-columns.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="D10024")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TABLE_FONT = Font(bold=True, size=11)
FIELD_FONT = Font(size=10)
SECTION_FONT = Font(bold=True, size=12)
REL_HEADER_FILL = PatternFill("solid", fgColor="2B2D42")
REL_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

# Table order left-to-right on the sheet
TABLE_ORDER = [
    "users",
    "categories",
    "brands",
    "products",
    "cities",
    "areas",
    "carts",
    "cart_items",
    "orders",
    "order_items",
    "order_logs",
    "reviews",
    "settings",
    "password_reset_tokens",
    "sessions",
    "personal_access_tokens",
]

ENTITIES = {
    "users": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("name", "VARCHAR", "", ""),
        ("email", "VARCHAR", "UNIQUE", ""),
        ("email_verified_at", "TIMESTAMP", "NULL", ""),
        ("password", "VARCHAR", "", ""),
        ("phone", "INT", "NULL", ""),
        ("role", "BOOLEAN", "DEFAULT 0", "0=customer, 1=admin"),
        ("google_id", "VARCHAR", "UNIQUE NULL", ""),
        ("avatar", "VARCHAR", "NULL", ""),
        ("remember_token", "VARCHAR", "NULL", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "password_reset_tokens": [
        ("email", "VARCHAR", "PK", ""),
        ("token", "VARCHAR", "", ""),
        ("created_at", "TIMESTAMP", "NULL", ""),
    ],
    "sessions": [
        ("id", "VARCHAR", "PK", ""),
        ("user_id", "BIGINT UNSIGNED", "FK→users NULL", ""),
        ("ip_address", "VARCHAR(45)", "NULL", ""),
        ("user_agent", "TEXT", "NULL", ""),
        ("payload", "LONGTEXT", "", ""),
        ("last_activity", "INT", "INDEX", ""),
    ],
    "categories": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("name", "VARCHAR", "UNIQUE", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "brands": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("name", "VARCHAR", "UNIQUE", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "products": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("category_id", "BIGINT UNSIGNED", "FK→categories", ""),
        ("brand_id", "BIGINT UNSIGNED", "FK→brands", ""),
        ("name", "VARCHAR", "", ""),
        ("description", "TEXT", "", ""),
        ("price", "INT", "", "stored in USD"),
        ("image", "VARCHAR", "", ""),
        ("stock", "INT", "", ""),
        ("sales", "INT", "DEFAULT 0", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "cities": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("name", "VARCHAR", "", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "areas": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("city_id", "BIGINT UNSIGNED", "FK→cities", ""),
        ("name", "VARCHAR", "", ""),
        ("fee", "DECIMAL(10,2)", "NULL DEFAULT 0", "USD"),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "carts": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("user_id", "BIGINT UNSIGNED", "FK→users", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "cart_items": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("cart_id", "BIGINT UNSIGNED", "FK→carts", ""),
        ("product_id", "BIGINT UNSIGNED", "FK→products", ""),
        ("quantity", "INT", "", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "orders": [
        ("id", "BIGINT UNSIGNED", "PK", "internal; customer sees order #"),
        ("user_id", "BIGINT UNSIGNED", "FK→users", ""),
        ("area_id", "BIGINT UNSIGNED", "FK→areas", ""),
        ("total_price", "DECIMAL(10,2)", "", "USD"),
        ("status", "ENUM", "", "pending,paid,shipped,delivered,canceled"),
        ("Note", "TEXT", "NULL", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "order_items": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("order_id", "BIGINT UNSIGNED", "FK→orders", ""),
        ("product_id", "BIGINT UNSIGNED", "FK→products", ""),
        ("price", "DECIMAL(10,2)", "", "USD at checkout"),
        ("quantity", "INT", "", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "order_logs": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("order_id", "BIGINT UNSIGNED", "FK→orders", ""),
        ("admin_id", "BIGINT UNSIGNED", "FK→users", ""),
        ("action", "VARCHAR", "", ""),
        ("old_status", "VARCHAR", "NULL", ""),
        ("new_status", "VARCHAR", "NULL", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "reviews": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("user_id", "BIGINT UNSIGNED", "FK→users", ""),
        ("product_id", "BIGINT UNSIGNED", "FK→products", ""),
        ("rating", "INT", "", "1-5"),
        ("comment", "TEXT", "", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "settings": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("key", "VARCHAR", "UNIQUE", "e.g. usd_to_sp_rate"),
        ("value", "TEXT", "", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
    "personal_access_tokens": [
        ("id", "BIGINT UNSIGNED", "PK", ""),
        ("tokenable_type", "VARCHAR", "", "polymorphic"),
        ("tokenable_id", "BIGINT UNSIGNED", "FK→users", ""),
        ("name", "VARCHAR", "", ""),
        ("token", "VARCHAR", "UNIQUE", ""),
        ("abilities", "TEXT", "NULL", ""),
        ("last_used_at", "TIMESTAMP", "NULL", ""),
        ("expires_at", "TIMESTAMP", "NULL", ""),
        ("created_at", "TIMESTAMP", "", ""),
        ("updated_at", "TIMESTAMP", "", ""),
    ],
}

# (parent, child, cardinality, fk_column, on_delete, note)
RELATIONSHIPS = [
    ("categories", "products", "1:N", "category_id", "CASCADE", ""),
    ("brands", "products", "1:N", "brand_id", "CASCADE", ""),
    ("users", "carts", "1:1", "user_id", "CASCADE", "one cart per user"),
    ("carts", "cart_items", "1:N", "cart_id", "CASCADE", ""),
    ("products", "cart_items", "1:N", "product_id", "CASCADE", ""),
    ("users", "orders", "1:N", "user_id", "CASCADE", ""),
    ("areas", "orders", "1:N", "area_id", "CASCADE", ""),
    ("cities", "areas", "1:N", "city_id", "CASCADE", ""),
    ("orders", "order_items", "1:N", "order_id", "CASCADE", ""),
    ("products", "order_items", "1:N", "product_id", "CASCADE", ""),
    ("orders", "order_logs", "1:N", "order_id", "CASCADE", ""),
    ("users", "order_logs", "1:N", "admin_id", "CASCADE", ""),
    ("users", "reviews", "1:N", "user_id", "CASCADE", ""),
    ("products", "reviews", "1:N", "product_id", "CASCADE", ""),
    ("users", "sessions", "1:N", "user_id", "SET NULL", ""),
    ("users", "personal_access_tokens", "1:N", "tokenable_id", "CASCADE", "via tokenable_type"),
]


def simplify_type(dtype):
    d = dtype.upper()
    mapping = [
        ("BIGINT UNSIGNED", "bigint"),
        ("BIGINT", "bigint"),
        ("VARCHAR", "string"),
        ("LONGTEXT", "longtext"),
        ("TEXT", "text"),
        ("TIMESTAMP", "timestamp"),
        ("DECIMAL", "decimal"),
        ("BOOLEAN", "boolean"),
        ("ENUM", "enum"),
        ("INT", "int"),
    ]
    for key, val in mapping:
        if key in d:
            if key == "VARCHAR" and "(" in dtype:
                return dtype.lower().replace("VARCHAR", "string")
            if key == "DECIMAL" and "(" in dtype:
                return dtype.lower()
            return val
    return dtype.lower()


def format_field(name, dtype, keys, note=""):
    type_str = simplify_type(dtype)
    parts = []

    if "PK" in keys:
        parts.append("PK")

    if "FK" in keys or "FK→" in keys:
        if "FK→" in keys:
            target = keys.split("FK→")[1].split()[0]
            parts.append(f"FK → {target}")
        else:
            parts.append("FK")

    if "UNIQUE" in keys:
        parts.append("unique")

    if "NULL" in keys and "PK" not in keys:
        parts.append("nullable")

    if "DEFAULT" in keys:
        default_part = keys.split("DEFAULT", 1)[1].strip()
        if default_part:
            parts.append(f"default {default_part.split()[0]}")

    if "INDEX" in keys:
        parts.append("indexed")

    line = f"{name} ({type_str})"
    if parts:
        line += " " + " ".join(parts)
    if note:
        line += f" — {note}"

    return line


def table_title(name):
    special = {
        "users": "User",
        "categories": "Category",
        "brands": "Brand",
        "products": "Product",
        "cities": "City",
        "areas": "Area",
        "carts": "Cart",
        "cart_items": "Cart_Item",
        "orders": "Order",
        "order_items": "Order_Item",
        "order_logs": "Order_Log",
        "reviews": "Review",
        "settings": "Setting",
        "password_reset_tokens": "Password_Reset_Token",
        "sessions": "Session",
        "personal_access_tokens": "Personal_Access_Token",
    }
    return special.get(name, name)


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "ERD"

    max_field_rows = max(len(ENTITIES[t]) for t in TABLE_ORDER)

    for col_idx, table in enumerate(TABLE_ORDER, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 28

        cell = ws.cell(row=1, column=col_idx, value=table_title(table))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_offset, (fname, ftype, fkeys, fnote) in enumerate(ENTITIES[table]):
            row = 2 + row_offset
            text = format_field(fname, ftype, fkeys, fnote)
            c = ws.cell(row=row, column=col_idx, value=text)
            c.font = FIELD_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")

    rel_start = 2 + max_field_rows + 2

    ws.merge_cells(
        start_row=rel_start,
        start_column=1,
        end_row=rel_start,
        end_column=6,
    )
    title_cell = ws.cell(row=rel_start, column=1, value="RELATIONSHIPS")
    title_cell.font = SECTION_FONT
    title_cell.alignment = Alignment(horizontal="left")

    headers = [
        "Parent (1 side)",
        "Child (N side)",
        "Cardinality",
        "FK column (on child)",
        "On delete",
        "Notes",
    ]
    header_row = rel_start + 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = REL_HEADER_FONT
        cell.fill = REL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            ws.column_dimensions[get_column_letter(col_idx)].width or 0,
            22 if col_idx <= 3 else 18,
        )

    for i, rel in enumerate(RELATIONSHIPS):
        row = header_row + 1 + i
        for col_idx, value in enumerate(rel, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = FIELD_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    note_row = header_row + 1 + len(RELATIONSHIPS) + 1
    ws.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=6,
    )
    ws.cell(
        row=note_row,
        column=1,
        value="No N:N relationships in this schema. All FKs are on the child (many) table. "
        "Customer order number is per-user sequence, not orders.id.",
    ).font = Font(italic=True, size=9)

    ws.freeze_panes = "A2"
    target = OUTPUT
    try:
        wb.save(target)
    except PermissionError:
        target = OUTPUT_ALT
        wb.save(target)
        print(f"Note: {OUTPUT.name} is locked (close Excel). Saved to alternate file.")
    print(f"Created: {target}")


if __name__ == "__main__":
    main()
